import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
import pdfkit
from jinja2 import Environment, FileSystemLoader
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Имя вакансии
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_average (float): Среднее значение оклада в рублях
        area_name (str): Город, в котором представлена вакансия
    """
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    def __init__(self, vacancy):
        """Инициализирует объект Vacancy

        Args:
            vacancy (dict): Словарь с информацией о вакансии
        """
        self.name = vacancy['name']
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    """Класс для парсинга csv и сбора необходимой информации о вакансиях.

        Attributes:
            filename (str): Имя csv файла
            vacancy_name (str): Имя выбранной вакансии
        """
    def __init__(self, filename, vacancy_name):
        self.filename = filename
        self.vacancy_name = vacancy_name

    @staticmethod
    def add(dict, key, value):
        """Добавляет выбранное значение к значению элемента по ключу или присваивает это значение, если по данному ключу не найдено элементов.

            Args:
                dict (dict): Словарь, по отношению к которому применяем операцию
                key (str): Ключ словаря, к значению которого применяем операцию
                value (str or float or int or double): Значение, которое необходимо добавить к сущестующему значению элемента по ключу или присвоить, если значения по ключу не существует
        """
        if key in dict:
            dict[key] += value
        else:
            dict[key] = value

    @staticmethod
    def avg(dict):
        """ Получить словарь, в котором все ключи такие же, как и в переданном, а значения - средние значения всех элементов по данному ключу из основного словаря.

            Args:
                dict (dict): Словарь, на основе которого будет возвращён новый

            Returns:
                dict: Новый словарь

        """
        new_dict = {}
        for key, values in dict.items():
            new_dict[key] = int(sum(values) / len(values))
        return new_dict

    def read_csv(self):
        """Чтение csv и формирование словарей с информацией о вакансиях.

            Returns:
                dict: Ленивый возврат словарей с информацией о вакансиях
        """

        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def get_info(self):
        """Получение информации для статистики.

            Returns:
                (dict,dict,dict,dict,dict,dict): 1.Зарплаты по годам, 2.Количество вакансий по годам, 3.Зарплаты по годам для выбранной профессии,
                    4.Количество вакансий по годам для выбранной профессии, 5.Зарплаты по городам, 6.Доля вакансий по городам
        """
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        count_of_vacancies = 0
        for vacancy_dictionary in self.read_csv():
            vacancy = Vacancy(vacancy_dictionary)
            self.add(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.add(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.add(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1
        years_count = dict([(key, len(value)) for key, value in salary.items()])
        years_count_vac = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])
        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            years_count_vac = dict([(key, 0) for key, value in years_count.items()])
        years_salary = self.avg(salary)
        years_salary_vac = self.avg(salary_of_vacancy_name)
        area_salary = self.avg(salary_city)
        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / count_of_vacancies, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        stats4.sort(key=lambda a: a[-1], reverse=True)
        area_count = stats4.copy()
        stats4 = dict(stats4)
        area_salary = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in area_salary.items()]))
        area_salary.sort(key=lambda a: a[-1], reverse=True)
        area_salary = dict(area_salary[:10])
        area_count = dict(area_count[:10])
        return years_salary, years_count, years_salary_vac, years_count_vac, area_salary, area_count


class Report:
    """Класс для формирования excel, pdf, png и текстового отчётов."""
    @staticmethod
    def generate_excel(data1, column_names, column_names2, data2):
        """Формирование excel отчёта в каталоге проекта.

            Args:
                data1 (list of dict): Список словарей с информацией для первого листа
                column_names (list of str): Названия столбцов первого листа
                column_names2 (list of str): Названия столбцов второго листа
                data2 (list of dict): Список словарей с информацией для второго листа
        """
        @staticmethod
        def auto_width(sheets):
            """Метод для установки автоматического размера ячейки на выбранных листах Workbook.

                Args:
                    sheets (list of Workbook): Список листов
            """

            for sheet in sheets:
                for sheet_column_cells in sheet.columns:
                    new_column_length = max(len(str(cell.value)) for cell in sheet_column_cells)
                    new_column_letter = (get_column_letter(sheet_column_cells[0].column))
                    if new_column_length > 0:
                        sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23

        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet("Статистика по городам")
        ws.title = 'Статистика по годам'
        border_style = Side(border_style="thin", color="000000")
        border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
        # Первый лист
        for i in range(len(column_names)):
            cell = ws.cell(row=1, column=i + 1, value=column_names[i])
            cell.font = Font(bold=True)
        for i, key in enumerate(data1[0]):
            ws.cell(row=i + 2, column=1, value=key)
            ws.cell(row=i + 2, column=2, value=data1[0][key])
            ws.cell(row=i + 2, column=3, value=data1[1][key])
            ws.cell(row=i + 2, column=4, value=data1[2][key])
            ws.cell(row=i + 2, column=5, value=data1[3][key])
        for column_cells in ws.columns:
            for cell in column_cells:
                cell.border = border
        # Второй лист
        lst = []
        lst.append(ws1.cell(row=1, column=1, value=column_names2[0]))
        lst.append(ws1.cell(row=1, column=2, value=column_names2[1]))
        lst.append(ws1.cell(row=1, column=4, value=column_names2[2]))
        lst.append(ws1.cell(row=1, column=5, value=column_names2[3]))
        for cell in lst:
            cell.font = Font(bold=True)
            cell.border = border
        for i, key in enumerate(data2[0]):
            ws1.cell(row=i+2, column=1, value=key)
            ws1.cell(row=i + 2, column=2, value=data2[0][key])
        for i, key in enumerate(data2[1]):
            ws1.cell(row=i+2, column=4, value=key)
            ws1.cell(row=i + 2, column=5, value=str('%.2f' % round(data2[1][key]*100, 2))+'%').number_format = '0.00%'
        for column_cells in ws1.columns:
            for cell in column_cells:
                if cell.column!=3:
                    cell.border = border
        auto_width([ws,ws1])
        wb.save('report.xlsx')

    @staticmethod
    def generate_image(data1, profession_name, data2, data3, data4):
        """Формирование png отчёта в каталоге проекта.

             Args:
                 data1 (list of dict): Cписок, включающий в себя: словарь с зарплатами по годам, словарь с зарплатами по годам
                    для выбранной вакансии
                 profession_name (str): Выбранная профессия
                 data2 (list of dict): Cписок, включающий в себя: словарь с количеством вакансий по годам, словарь с
                    количеством вакансий по годам для выбранной вакансии
                 data3 (dict): Словарь с уровнями зарплат по городам
                 data4 (dict): Словарь с долями вакансий по городам
         """
        fig,axs = plt.subplots(2, 2)
        #Первый график
        width = 0.4
        plt.rcParams.update({'font.size': 8})
        axs[0, 0].bar([x - width * 0.5 for x in list(data1[0].keys())], data1[0].values(), width, label='средняя з/п')
        axs[0, 0].bar([x + width * 0.5 for x in list(data1[1].keys())], data1[1].values(), width,
                label=f'з/п {profession_name}')
        axs[0, 0].legend()
        axs[0, 0].set_title("Уровень зарплат по годам", fontsize=12)
        axs[0, 0].grid(axis='y')
        plt.sca(axs[0, 0])
        plt.xticks(np.arange(min(list(data1[0].keys())), max(list(data1[0].keys()))+1, 1.0),rotation=90)
        # Второй график
        axs[0, 1].bar([x - width * 0.5 for x in list(data2[0].keys())], data2[0].values(), width,
                 label='Количество вакансий')
        axs[0, 1].bar([x + width * 0.5 for x in list(data2[1].keys())], data2[1].values(), width,
                 label=f'Количество вакансий \n{profession_name}')
        axs[0, 1].grid(axis='y')
        plt.sca(axs[0, 1])
        plt.xticks(np.arange(min(list(data1[1].keys())), max(list(data1[1].keys()))+1, 1.0), rotation=90)
        axs[0, 1].legend()
        axs[0, 1].set_title('Количество вакансий по годам', fontsize=12)
        # Третий график
        width = 0.8
        keys = list(data3.keys())
        for i, key in enumerate(keys):
            keys[i] = keys[i].replace(" ", '\n')
            keys[i] = keys[i].replace('-', '-\n')
        keys.reverse()
        values = list(data3.values())
        values.reverse()
        axs[1, 0].barh(keys, values, width, label='Количество вакансий')
        axs[1, 0].set_title('Уровень зарплат по городам', fontsize=12)
        axs[1, 0].grid(axis='x')
        plt.sca(axs[1, 0])
        plt.yticks(fontsize=6)
        plt.xticks(np.arange(0, max(values) + 20000, 20000))
        # Четвёртый график
        share = 1
        for key in data4:
            share -= data4[key]
        data4['Другие'] = share
        axs[1, 1].pie(data4.values(), labels=data4.keys(), startangle=450, textprops={'fontsize': 6})
        axs[1, 1].set_title('Доля вакансий по городам', fontsize=12)
        # Вывод графика и сохранение в .png
        fig.tight_layout()
        plt.show()
        fig.savefig('graph.png')

    @staticmethod
    def generate_pdf(profession_name, data1, column_names1, data2, column_names2, data3, column_names3):
        """Формирование png отчёта в каталоге проекта.

             Args:
                 profession_name (str): Выбранная профессия
                 data1 (list of dict): Cписок, включающий в себя: словарь с зарплатами по годам, словарь с зарплатами по годам
                    для выбранной вакансии, словарь с количеством вакансий по годам, словарь с количеством вакансий по годам
                    для выбранной вакансии
                 column_names1 (list of str): Названия столбцов первой таблицы
                 data2 (dict): Словарь с уровнями зарплат по городам
                 data3 (dict): Словарь с долями зарплат по городам
        """
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        table1_columns = ""
        table1_rows = ""
        table2_columns = ""
        table2_rows = ""
        table3_columns = ""
        table3_rows = ""
        for i in range(0,len(column_names1)):
            table1_columns += '<th>' + column_names1[i] + '</th>'
        for i in range(0, len(data1[0])):
            table1_rows += '<tr>' + f'<td>{list(data1[0].keys())[i]}</td>' + f'<td><center>{list(data1[0].values())[i]}</center></td>' + \
                           f'<td><center>{list(data1[1].values())[i]}</center></td>' + f'<td><center>{list(data1[2].values())[i]}</center></td>' + \
                           f'<td><center>{list(data1[3].values())[i]}</center></td>' + '</tr>'
        for i in range(0, len(column_names2)):
            table2_columns += '<th>' + column_names2[i] + '</th>'
        for i in range(0, len(data2)):
            table2_rows += '<tr>' + f'<td><center>{list(data2.keys())[i]}</center></td>' + f'<td><center>{list(data2.values())[i]}</center></td>' + '</tr>'
        for i in range(0, len(column_names3)):
            table3_columns += '<th>' + column_names3[i] + '</th>'
        for i in range(0, len(data3)):
            table3_rows += '<tr>' + f'<td><center>{list(data3.keys())[i]}</center></td>' + f'<td><center>{list(data3.values())[i]}</center></td>' + '</tr>'
        pdf_template = template.render({'profession_name': profession_name, 'image_path':os.path.abspath('graph.png'),
                                        'table1_columns':table1_columns, 'table1_rows':table1_rows, 'table2_columns':table2_columns, 'table2_rows':table2_rows,
                                        'table3_columns':table3_columns, 'table3_rows':table3_rows})
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    @staticmethod
    def print_report(years_salary, years_count, years_salary_vac, years_count_vac, area_salary, area_count):
        """Формирование и вывод текстового отчёта.

             Args:
                 years_salary (dict): Словарь с зарплатами по годам
                 years_count (dict): Словарь с количеством вакансий по годам
                 years_salary_vac (dict): Словарь с зарплатами по годам для выбранной профессии
                 years_count_vac (dict): Словарь с количеством вакансий по годам для выбранной профессии
                 area_salary (dict): Словарь с уровнями зарплат по городам
                 area_count (dict): Словарь с долями вакансий по городам
        """
        print("Динамика уровня зарплат по годам: " + str(years_salary))
        print("Динамика количества вакансий по годам: " + str(years_count))
        print("Динамика уровня зарплат по годам для выбранной профессии: " + str(years_salary_vac))
        print("Динамика количества вакансий по годам для выбранной профессии: " + str(years_count_vac))
        print("Уровень зарплат по городам (в порядке убывания): " + str(area_salary))
        print("Доля вакансий по городам (в порядке убывания): " + str(area_count))


class Program:
    """Класс для инициализации необходимых для работы программы данных.

        Attributes:
            program_action (str): Выбор необходимого алгоритма (Статистика - для формирования png, excel и pdf отчётов,
                Вакансии - для формирования текстового отчёта)
            report (Report): Экземпляр класса Report
            dataset (DataSet): Экземпляр класса DataSet
            years_salary_dic (dict): Словарь с зарплатами по годам
            years_count_dic (dict): Словарь с количеством вакансий по годам
            years_salary_vac_dic (dict): Словарь с зарплатами по годам для выбранной профессии
            years_count_vac_dic (dict): Словарь с количеством вакансий по годам для выбранной профессии
            area_salary_dic (dict): Словарь с уровнями зарплат по городам
            area_count_dic (dict): Словарь с долями вакансий по городам

    """
    def __init__(self):
        """Инициализирует объект Program, присваивает переменным необходимые данные на основе пользовательского ввода."""
        self.program_action = input('Введите данные для печати: ')
        self.report = Report()
        self.dataset = DataSet(input("Введите название файла: "), input("Введите название профессии: "))
        self.years_salary_dic, self.years_count_dic, self.years_salary_vac_dic, self.years_count_vac_dic, self.area_salary_dic, self.area_count_dic = self.dataset.get_info()

    @staticmethod
    def as_text(value):
        """Конвертирует значение какого-либо типа в строку

            Returns:
                str: Строковое представление значения
        """
        if value is None:
            return ""
        return str(value)

    def run(self):
        """Запускает необходимые для формирования выбранного типа отчёта операции."""
        if self.program_action == 'Вакансии':
            self.report.print_report(self.years_salary_dic, self.years_count_dic, self.years_salary_vac_dic,
                                     self.years_count_vac_dic, self.area_salary_dic, self.area_count_dic)
        elif self.program_action == 'Статистика':
            first_sheet_data = [self.years_salary_dic, self.years_salary_vac_dic, self.years_count_dic, self.years_count_vac_dic]
            first_table_data = [self.years_salary_dic, self.years_salary_vac_dic, self.years_count_dic, self.years_count_vac_dic]
            second_sheet_data = [self.area_salary_dic, self.area_count_dic]
            self.report.generate_excel(first_sheet_data,
                        ['Год', 'Средняя зарплата', 'Средняя зарплата - {0}'.format(self.dataset.vacancy_name), 'Количество вакансий',
                        'Количество вакансий - {0}'.format(self.dataset.vacancy_name)], ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий'],second_sheet_data)
            self.report.generate_image([self.years_salary_dic, self.years_salary_vac_dic], self.dataset.vacancy_name, [self.years_count_dic, self.years_count_vac_dic]
                        , self.area_salary_dic, self.area_count_dic)
            table3_data = {}
            for i in range(0, len(self.area_count_dic)-1):
                value = str('%.2f' % round(list(self.area_count_dic.values())[i]*100, 2)+'%')
                table3_data[list(self.area_count_dic.keys())[i]] = value
            self.report.generate_pdf(self.dataset.vacancy_name,first_table_data,['Год', 'Средняя зарплата', 'Средняя зарплата - {0}'.format(self.dataset.vacancy_name), 'Количество вакансий', 'Количество вакансий - {0}'.format(self.dataset.vacancy_name)],
                                     self.area_salary_dic,['Город', 'Уровень зарплат'], table3_data, ['Город', 'Доля вакансий'])


program = Program()
program.run()


